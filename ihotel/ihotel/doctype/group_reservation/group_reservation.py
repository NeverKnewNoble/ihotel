# Copyright (c) 2026, Noble and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils import getdate, date_diff, cint, flt


class GroupReservation(Document):
	def validate(self):
		self.validate_dates()
		self.calculate_days()
		self.calculate_revenue()
		self.calculate_deposit()
		self.validate_rooming_list()

	def calculate_deposit(self):
		"""Auto-derive deposit_required from total_room_revenue × deposit_percent."""
		if self.deposit_percent is None or self.deposit_percent == "":
			self.deposit_percent = 30
		self.deposit_required = round(
			flt(self.total_room_revenue) * flt(self.deposit_percent) / 100.0, 2
		)
		self.deposit_balance = round(
			flt(self.deposit_required) - flt(self.deposit_amount), 2
		)

	def validate_dates(self):
		if self.check_in_date and self.check_out_date:
			if getdate(self.check_in_date) >= getdate(self.check_out_date):
				frappe.throw(_("Check-in date must be before check-out date"))

		if self.cutoff_date and self.check_in_date:
			if getdate(self.cutoff_date) > getdate(self.check_in_date):
				frappe.throw(_("Room block cutoff date cannot be after check-in date"))

	def calculate_days(self):
		if self.check_in_date and self.check_out_date:
			self.days = date_diff(self.check_out_date, self.check_in_date)

	def calculate_revenue(self):
		if self.days and self.no_of_rooms and self.rate_per_night:
			self.total_room_revenue = flt(self.days) * flt(self.no_of_rooms) * flt(self.rate_per_night)

	def validate_rooming_list(self):
		if self.rooming_list and len(self.rooming_list) > cint(self.no_of_rooms or 0):
			frappe.throw(_(
				"Rooming list has {0} entries but only {1} room(s) are reserved."
			).format(len(self.rooming_list), self.no_of_rooms))

	def on_submit(self):
		if self.status == "Tentative":
			self.db_set("status", "Confirmed")

	def on_update(self):
		"""Cascade cancellation to all generated child reservations."""
		if self.status == "Cancelled":
			child_reservations = frappe.get_all(
				"Reservation",
				filters={"group_reservation": self.name, "status": ["!=", "cancelled"]},
				pluck="name",
			)
			for res_name in child_reservations:
				frappe.db.set_value("Reservation", res_name, "status", "cancelled",
					update_modified=False)
			if child_reservations:
				frappe.msgprint(
					_("{0} reservation(s) cancelled along with this group.").format(
						len(child_reservations)
					),
					indicator="orange",
					alert=True,
				)


@frappe.whitelist()
def generate_reservations(group_reservation_name):
	group = frappe.get_doc("Group Reservation", group_reservation_name)

	if not group.no_of_rooms or cint(group.no_of_rooms) <= 0:
		frappe.throw(_("Please set the number of rooms"))

	# Check how many reservations already exist for this group
	existing = frappe.db.count("Reservation", {"group_reservation": group.name})
	remaining = cint(group.no_of_rooms) - existing

	if remaining <= 0:
		frappe.throw(_("All {0} reservations have already been generated").format(group.no_of_rooms))

	created = []
	for _i in range(remaining):
		reservation = frappe.get_doc({
			"doctype": "Reservation",
			"group_reservation": group.name,
			"check_in_date": group.check_in_date,
			"check_out_date": group.check_out_date,
			"check_in_time": group.check_in_time,
			"check_out_time": group.check_out_time,
			"room_type": group.room_type,
			"rate_type": group.rate_type,
			"rent": group.rate_per_night,
			"business_source_category": group.business_source_category,
			"full_name": group.full_name,
			"company": group.company,
			"email_address": group.email,
			"phone_number": group.phone_number,
			"city": group.city,
			"state": group.state,
			"country": group.country,
			"postal_code": group.zip_code,
			"status": "pending",
		})
		reservation.insert(ignore_permissions=True)
		created.append(reservation.name)

	frappe.msgprint(
		_("{0} reservations created successfully").format(len(created)),
		indicator="green",
		alert=True,
	)

	return created


@frappe.whitelist()
def get_group_party_details(group_name):
	if not group_name:
		return {}
	return _get_guest_details(group_name)


def _get_guest_details(guest_name):
	guest = frappe.get_doc("Guest", guest_name)
	address_parts = [guest.address_line_1, guest.address_line_2]
	address = ", ".join([part for part in address_parts if part])

	return {
		"full_name": guest.guest_name or guest.name,
		"company": guest.customer or "",
		"email": guest.email or "",
		"phone_number": guest.phone or "",
		"address": address,
		"city": guest.city or "",
		"state": guest.state or "",
		"country": guest.nationality or "",
		"zip_code": guest.postal_code or "",
	}


# ---------------------------------------------------------------------------
# Rooming List — Excel template download + bulk upload
# ---------------------------------------------------------------------------
ROOMING_LIST_HEADERS = ["Guest Name", "Room Type", "Room", "Adults", "Children", "Special Requests"]
ROOMING_LIST_FIELDS  = ["guest_name", "room_type", "room", "adults", "children", "special_requests"]


@frappe.whitelist()
def download_rooming_list_template(group_reservation_name=None):
	"""Stream an XLSX template for the rooming list with one example row."""
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill
	from io import BytesIO

	wb = Workbook()
	ws = wb.active
	ws.title = "Rooming List"

	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

	for col_idx, header in enumerate(ROOMING_LIST_HEADERS, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = header_font
		cell.fill = header_fill

	# Example row — pull the group's default room type if available
	default_rt = ""
	if group_reservation_name:
		default_rt = frappe.db.get_value("Group Reservation", group_reservation_name, "room_type") or ""
	ws.append(["John Doe", default_rt, "", 1, 0, ""])

	# Reasonable column widths
	for col_idx, _h in enumerate(ROOMING_LIST_HEADERS, start=1):
		ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

	buf = BytesIO()
	wb.save(buf)
	buf.seek(0)

	frappe.local.response.filename = f"rooming_list_template_{group_reservation_name or 'group'}.xlsx"
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "binary"


@frappe.whitelist()
def upload_rooming_list(group_reservation_name, file_url):
	"""Append rows from an uploaded XLSX/XLS to the group's rooming_list child table.

	Behaviour:
	- Append (does not replace existing rows).
	- Skip rows that would push the total above no_of_rooms; warn for each.
	- Skip rows with invalid Room Type; warn.
	- Skip rows with a Room that doesn't belong to the row's Room Type; warn.
	- Trim whitespace on text fields. Default adults to 1 if blank.
	"""
	from openpyxl import load_workbook
	import os

	doc = frappe.get_doc("Group Reservation", group_reservation_name)
	doc.check_permission("write")

	# Resolve file path
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	file_path = file_doc.get_full_path()
	if not os.path.exists(file_path):
		frappe.throw(_("Uploaded file not found on server."))

	wb = load_workbook(file_path, data_only=True, read_only=True)
	ws = wb.active

	rows = list(ws.iter_rows(values_only=True))
	if not rows:
		return {"imported": 0, "total": 0, "warnings": [_("File is empty.")]}

	# Detect header row
	header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
	expected = [h.lower() for h in ROOMING_LIST_HEADERS]
	if header[: len(expected)] != expected:
		return {
			"imported": 0,
			"total": 0,
			"warnings": [_("Header row does not match the template. Expected: {0}").format(", ".join(ROOMING_LIST_HEADERS))],
		}

	data_rows = rows[1:]
	# Filter out completely-blank rows
	data_rows = [r for r in data_rows if any(c not in (None, "") for c in r)]

	capacity = cint(doc.no_of_rooms or 0)
	already   = len(doc.rooming_list or [])
	remaining = max(capacity - already, 0)
	default_rt = doc.room_type

	warnings = []
	imported = 0

	for i, raw in enumerate(data_rows, start=2):  # row 1 is header
		if remaining <= 0:
			warnings.append(_("Row {0}: skipped — group capacity of {1} room(s) reached.").format(i, capacity))
			continue

		# Pad row to expected length
		raw = list(raw) + [None] * (len(ROOMING_LIST_FIELDS) - len(raw))
		guest_name, room_type, room, adults, children, special_requests = raw[: len(ROOMING_LIST_FIELDS)]

		guest_name = (str(guest_name).strip() if guest_name else "")
		room_type  = (str(room_type).strip() if room_type else "") or default_rt or ""
		room       = (str(room).strip() if room else "")
		adults_val   = cint(adults) if adults not in (None, "") else 1
		children_val = cint(children) if children not in (None, "") else 0
		special_requests = (str(special_requests).strip() if special_requests else "")

		# Infer room_type from the room when missing
		if not room_type and room:
			room_type = frappe.db.get_value("Room", room, "room_type") or ""

		# Validate room_type
		if room_type and not frappe.db.exists("Room Type", room_type):
			warnings.append(_("Row {0}: skipped — Room Type '{1}' does not exist.").format(i, room_type))
			continue

		# Validate room ↔ room_type
		if room:
			if not frappe.db.exists("Room", room):
				warnings.append(_("Row {0}: skipped — Room '{1}' does not exist.").format(i, room))
				continue
			actual_rt = frappe.db.get_value("Room", room, "room_type")
			if room_type and actual_rt != room_type:
				warnings.append(_(
					"Row {0}: skipped — Room '{1}' is type '{2}', not '{3}'."
				).format(i, room, actual_rt, room_type))
				continue

		doc.append("rooming_list", {
			"guest_name": guest_name,
			"room_type": room_type,
			"room": room or None,
			"adults": adults_val,
			"children": children_val,
			"special_requests": special_requests,
		})
		imported += 1
		remaining -= 1

	if imported:
		doc.save(ignore_permissions=False)

	return {
		"imported": imported,
		"total": len(data_rows),
		"warnings": warnings,
	}
